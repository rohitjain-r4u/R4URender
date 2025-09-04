import os

# Req_App.py
from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, send_file
from flask_wtf.csrf import CSRFProtect
import psycopg2
import psycopg2.extras
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime
from contextlib import contextmanager
from itsdangerous import URLSafeTimedSerializer
import logging
import json
import io
import openpyxl
import re
from export import export_bp
from pagination import Paginator, sanitize_page_params
from AllCandidates import all_candidates_bp   # import the blueprint
from dashboard_routes import dashboard_bp
from Reports import reports_bp
from recruiter_performance import recruiter_perf_bp
from datetime import datetime, timedelta
from flask import jsonify

app = Flask(__name__)

from flask_mail import Mail, Message

# Email config using environment variables
EMAIL_USER = os.getenv('GMAIL_USER', 'hr@recruiteforu.com')
EMAIL_PASS = os.getenv('GMAIL_APP_PASSWORD', 'rbdh wemz vbyq ffyx')

app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USERNAME'] = EMAIL_USER
app.config['MAIL_PASSWORD'] = EMAIL_PASS
app.config['MAIL_DEFAULT_SENDER'] = ('Recruitment Portal', EMAIL_USER)

mail = Mail(app)

app.secret_key = 'ueueuweujfjdjsdjsdsjdajsdajsdlajsdlajdlajd'  # Change this to a secure secret in production
csrf = CSRFProtect(app)

from flask_wtf.csrf import generate_csrf

from flask_wtf.csrf import CSRFProtect, generate_csrf
app.register_blueprint(all_candidates_bp)     # register it
app.register_blueprint(dashboard_bp)
app.register_blueprint(reports_bp)
app.register_blueprint(recruiter_perf_bp)

@app.context_processor
def inject_csrf_token():
    return dict(csrf_token=generate_csrf)
app.config['TEMPLATES_AUTO_RELOAD'] = True

# DB config - env-driven (supports DATABASE_URL or individual DB_* vars)
DB_CONFIG = {}
DATABASE_URL = os.getenv('DATABASE_URL')
if DATABASE_URL:
    DB_CONFIG['dsn'] = DATABASE_URL
else:
    DB_CONFIG = {
        'dbname': os.getenv('DB_NAME', 'job_portal'),
        'user': os.getenv('DB_USER', 'postgres'),
        'password': os.getenv('DB_PASSWORD', ''),
        'host': os.getenv('DB_HOST', 'localhost'),
        'port': int(os.getenv('DB_PORT', '5432')),
    }
from import_routes import import_bp
app.register_blueprint(import_bp, url_prefix="/candidates/import")
app.register_blueprint(export_bp)

# Token serializer for password reset links
ts = URLSafeTimedSerializer(app.secret_key)

# Logger
logger = logging.getLogger('req_app')
logging.basicConfig(level=logging.INFO)


@contextmanager
def get_db_cursor():
    conn = None
    cur = None
    try:
        conn = psycopg2.connect(DB_CONFIG['dsn']) if 'dsn' in DB_CONFIG else psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        yield conn, cur
    except Exception:
        app.logger.exception("DB connection or query error")
        raise
    finally:
        if cur:
            try:
                cur.close()
            except Exception:
                pass
        if conn:
            try:
                conn.close()
            except Exception:
                pass


# small helper: normalize JSON/list fields to Python lists for templates
def normalize_list_field(value):
    """
    Accepts value which may be:
    - Python list -> return as-is
    - JSON string (e.g. '["a","b"]') -> parsed list
    - Plain string -> wrap in single-element list (unless empty)
    - None -> return []
    - other scalar -> wrap as list
    """
    if value is None:
        return []
    if isinstance(value, list):
        return value
    if isinstance(value, (dict,)):
        # unusual: dict -> convert to JSON string inside list for display safety
        try:
            return [json.dumps(value, ensure_ascii=False)]
        except Exception:
            return [str(value)]
    if isinstance(value, str):
        v = value.strip()
        if not v:
            return []
        # try parse JSON
        try:
            parsed = json.loads(value)
            if isinstance(parsed, list):
                return parsed
            else:
                # parsed but not a list - wrap
                return [parsed]
        except Exception:
            # not JSON - return single-element list
            return [value]
    # fallback: wrap whatever it is
    return [value]


# --- Minimal validators# --- Minimal validators (relaxed per request) ---
def validate_requirement_form(form):
    """
    Minimal validation: only enforce required fields (client_name, requirement_name).
    All other fields accept special characters / free text.
    """
    errors = []
    data = {}

    client_name = form.get('client_name', '').strip()
    if not client_name:
        errors.append('Client name is required.')
    data['client_name'] = client_name

    requirement_name = form.get('requirement_name', '').strip()
    if not requirement_name:
        errors.append('Requirement name is required.')
    data['requirement_name'] = requirement_name

    data['experience'] = form.get('experience', '').strip()
    data['mandatory_skills'] = form.get('mandatory_skills', '').strip()
    data['job_locations'] = form.get('job_locations', '').strip()
    data['remote'] = bool(form.get('remote'))  # checkbox -> boolean
    data['budget'] = form.get('budget', '').strip()
    data['job_description'] = form.get('job_description', '').strip()
    data['job_d_th_d'] = form.get('job_d_th_d', '').strip()

    # NEW FIELDS
    data['client_linkedin_profile'] = form.get('client_linkedin_profile', '').strip()
    data['client_brief_description'] = form.get('client_brief_description', '').strip()

    # Assigned To: MULTI-SELECT CHECKBOXES -> comma-separated usernames
    assigned_list = form.getlist('assigned_to')  # e.g., ['alice','bob']
    assigned_list = [u.strip() for u in assigned_list if u and u.strip()]
    data['assigned_to'] = ", ".join(sorted(set(assigned_list))) if assigned_list else ""

    # Status defaults to Active; "Open" removed
    status = form.get('status', 'Active')
    if status not in ('Active', 'Hold', 'Closed'):
        status = 'Active'
    data['status'] = status

    return data, errors


# -----------------------
# Candidate validation helper
# -----------------------
def validate_candidate_form(form):
    """
    Returns (data_dict, errors_list).
    Expects multiple phones/emails submitted with name="phones" / name="emails" (form.getlist).
    """
    errors = []
    data = {}

    data['application_date'] = form.get('application_date') or None
    data['job_title'] = (form.get('job_title') or '').strip()
    data['candidate_name'] = (form.get('candidate_name') or '').strip()
    if not data['candidate_name']:
        errors.append('Candidate name is required.')

    data['current_company'] = (form.get('current_company') or '').strip()
    data['total_experience'] = (form.get('total_experience') or '').strip()

    phones = form.getlist('phones') or []
    phones = [p.strip() for p in phones if p and p.strip()]
    data['phones'] = phones

    emails = form.getlist('emails') or []
    emails = [e.strip() for e in emails if e and e.strip()]
    data['emails'] = emails

    data['notice_period'] = (form.get('notice_period') or '').strip()
    data['current_location'] = (form.get('current_location') or '').strip()
    data['preferred_locations'] = (form.get('preferred_locations') or '').strip()

    def _num_or_none(x):
        try:
            return float(x) if x is not None and x != '' else None
        except:
            return None

    data['ctc_current'] = _num_or_none(form.get('ctc_current'))
    data['ectc'] = _num_or_none(form.get('ectc'))
    data['key_skills'] = (form.get('key_skills') or '').strip()
    data['education'] = (form.get('education') or '').strip()
    data['post_graduation'] = (form.get('post_graduation') or '').strip()
    data['pf_docs_confirm'] = True if form.get('pf_docs_confirm') in ('1', 'true', 'on', 'yes') else False
    data['notice_period_details'] = (form.get('notice_period_details') or '').strip()
    data['current_ctc_lpa'] = _num_or_none(form.get('current_ctc_lpa'))
    data['expected_ctc_lpa'] = _num_or_none(form.get('expected_ctc_lpa'))
    try:
        data['employee_size'] = int(form.get('employee_size')) if form.get('employee_size') else None
    except:
        data['employee_size'] = None
    data['companies_worked'] = (form.get('companies_worked') or '').strip()
    data['calling_status'] = (form.get('calling_status') or '').strip()
    data['profile_status'] = (form.get('profile_status') or '').strip()
    data['comments'] = (form.get('comments') or '').strip()

    data['interview_date'] = (form.get('interview_date') or None)
    data['interview_time'] = (form.get('interview_time') or None)

    return data, errors


# -----------------------
# Routes
# -----------------------
@app.route('/')
def dashboard():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    user_role = session.get('role')
    return render_template('dashboard.html', role=user_role)


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        login_input = request.form.get('login')
        password = request.form.get('password')
        remember = True if request.form.get('remember') else False

        if not login_input or not password:
            flash('Please fill in all fields', 'danger')
            return redirect(url_for('login'))

        try:
            with get_db_cursor() as (conn, cur):
                is_email = '@' in login_input
                # fetch username too so recruiter filtering by username works
                if is_email:
                    cur.execute("SELECT id, role, password_hash, username FROM users WHERE email = %s", (login_input,))
                else:
                    cur.execute("SELECT id, role, password_hash, username FROM users WHERE username = %s", (login_input,))
                user = cur.fetchone()
                if user and check_password_hash(user['password_hash'], password):
                    session['user_id'] = user['id']
                    session['role'] = user['role']
                    session['username'] = user.get('username')  # used for recruiter filtering
                    if remember:
                        session.permanent = True
                    flash('Login successful!', 'success')
                    return redirect(url_for('dashboard'))
                else:
                    flash('Invalid login credentials', 'danger')
        except Exception:
            app.logger.exception("Login error")
            flash('Login error', 'danger')

    return render_template('login.html')


@app.route('/logout')
def logout():
    session.clear()
    flash('You have been logged out', 'info')
    return redirect(url_for('login'))



@app.route('/requirements')
def requirements():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    # pagination and filtering
    try:
        page = int(request.args.get('page', 1))
    except Exception:
        page = 1
    try:
        per_page = int(request.args.get('per_page', 20))
    except Exception:
        per_page = 20

    sort_by = request.args.get('sort_by', 'added_date')
    sort_dir = request.args.get('sort_dir', 'desc').lower()
    status_filter = request.args.get('status')
    client_filter = request.args.get('client')               # search by client name
    requirement_filter = request.args.get('requirement')     # search by requirement name
    assigned_filter = request.args.get('assigned_to')       # optional

    allowed_sort_columns = {
        'added_date': 'r.added_date',
        'experience': 'r.experience',
        'budget': 'r.budget',
        'id': 'r.id',
        'client_name': 'r.client_name',
        'requirement_name': 'r.requirement_name',
        'assigned_to': 'r.assigned_to',   # sort by CSV usernames (TEXT)
        'status': 'r.status'              # sort by Status
    }
    order_column = allowed_sort_columns.get(sort_by, 'r.added_date')
    order_dir = 'DESC' if sort_dir == 'desc' else 'ASC'
    offset = (page - 1) * per_page

    requirements = []
    total = 0

    try:
        with get_db_cursor() as (conn, cur):
            base_query = """
                SELECT r.*
                FROM requirements r
            """
            where_clauses = []
            params = []

            # Role-restricted view for recruiters: show only records where their username appears in CSV
            if session.get('role') == 'recruiter':
                me_username = session.get('username')
                if not me_username:
                    cur.execute("SELECT username FROM users WHERE id = %s", (session['user_id'],))
                    row = cur.fetchone()
                    me_username = row.get('username') if row else None
                    session['username'] = me_username
                if me_username:
                    where_clauses.append("r.assigned_to ILIKE %s")
                    params.append(f"%{me_username}%")

            if status_filter:
                where_clauses.append("r.status = %s")
                params.append(status_filter)

            if client_filter:
                where_clauses.append("r.client_name ILIKE %s")
                params.append(f"%{client_filter}%")

            if requirement_filter:
                where_clauses.append("r.requirement_name ILIKE %s")
                params.append(f"%{requirement_filter}%")

            if assigned_filter:
                # allow filtering by assigned username fragment
                where_clauses.append("r.assigned_to ILIKE %s")
                params.append(f"%{assigned_filter}%")

            if where_clauses:
                base_query += " WHERE " + " AND ".join(where_clauses)

            # final query with ordering and pagination
            final_query = f"{base_query} ORDER BY {order_column} {order_dir} LIMIT %s OFFSET %s"
            query_params = tuple(params + [per_page, offset])
            cur.execute(final_query, query_params)
            requirements = cur.fetchall()

            # count using the same WHERE clauses
            count_query = "SELECT COUNT(*) as total FROM requirements r"
            if where_clauses:
                count_query += " WHERE " + " AND ".join(where_clauses)
                count_params = tuple(params)
                cur.execute(count_query, count_params)
            else:
                cur.execute(count_query)
            total_row = cur.fetchone()
            total = total_row.get('total', 0) if total_row else 0

    except Exception:
        app.logger.exception("Error fetching requirements")
        flash("Error fetching requirements", "danger")
        requirements = []
        total = 0

    return render_template(
        'requirements.html',
        requirements=requirements,
        role=session.get('role'),
        page=page,
        per_page=per_page,
        total=total,
        sort_by=sort_by,
        sort_dir=sort_dir,
        client_filter=client_filter or '',
        requirement_filter=requirement_filter or '',
        status_filter=status_filter or '',
        assigned_filter=assigned_filter or ''
    )

from flask import request, redirect, url_for, flash, render_template, session
from werkzeug.security import generate_password_hash  # keep this at top with imports


def send_requirement_email(requirement, assigned_users):
    """Send requirement details to recruiters/admins."""
    if not assigned_users:
        return

    subject = f"New Requirement Assigned: {requirement['requirement_name']}"
    body = f"""
    Dear Recruiter,

    A new requirement has been assigned to you.

    Client: {requirement['client_name']}
    Requirement: {requirement['requirement_name']}
    Experience: {requirement['experience']}
    Skills: {requirement['mandatory_skills']}
    Location: {requirement['job_locations']}
    Budget: {requirement['budget']}
    Description: {requirement['job_description']}
    Status: {requirement['status']}

    Please log in to the portal for full details.
    """

    with get_db_cursor() as (conn, cur):
        cur.execute("SELECT email FROM users WHERE username = ANY(%s)", (assigned_users,))
        rows = cur.fetchall()
        emails = [r['email'] for r in rows if r.get('email')]

    if not emails:
        app.logger.warning("No email IDs found for assigned users: %s", assigned_users)
        return

    try:
        msg = Message(subject=subject, recipients=emails, body=body)
        mail.send(msg)
        app.logger.info("Requirement emails sent to: %s", emails)
    except Exception:
        app.logger.exception("Failed to send requirement email")


# --- USER MANAGEMENT ---

@app.route('/add_requirement', methods=['GET', 'POST'])
def add_requirement():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    recruiters = []
    try:
        with get_db_cursor() as (conn, cur):
            # Show all recruiters or admins (active filter can be added if your users table has such a flag)
            cur.execute("SELECT username FROM users WHERE role IN ('recruiter','admin') ORDER BY username ASC")
            recruiters = [r['username'] for r in cur.fetchall()]

            if request.method == 'POST':
                data, errors = validate_requirement_form(request.form)
                if errors:
                    for e in errors:
                        flash(e, 'danger')
                    return render_template('add_requirement.html', recruiters=recruiters)

                cur.execute("""
                    INSERT INTO requirements (
                        client_name, requirement_name, experience, mandatory_skills,
                        job_locations, remote, budget, job_description, job_d_th_d,
                        client_linkedin_profile, client_brief_description,
                        assigned_to, status, added_date
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    RETURNING id
                """, (
                    data['client_name'],
                    data['requirement_name'],
                    data['experience'],
                    data['mandatory_skills'],
                    data['job_locations'],
                    data['remote'],
                    data['budget'],
                    data['job_description'],
                    data['job_d_th_d'],
                    data['client_linkedin_profile'],
                    data['client_brief_description'],
                    data['assigned_to'],  # CSV usernames
                    data['status'],
                    datetime.now()
                ))
                new_row = cur.fetchone()
                new_id = new_row['id'] if new_row else None
                
                conn.commit()
                flash('Requirement added successfully!', 'success')

                # Fetch inserted requirement to include details
                cur.execute("SELECT * FROM requirements WHERE id = %s", (new_id,))
                req = cur.fetchone()
                assigned_users = [u.strip() for u in (req['assigned_to'] or '').split(',') if u.strip()]
                send_requirement_email(req, assigned_users)

                return redirect(url_for('requirements'))


    except Exception:
        app.logger.exception("Error adding requirement")
        flash('Error adding requirement', 'danger')

    return render_template('add_requirement.html', recruiters=recruiters)




@app.route('/requirement_detail/<int:req_id>', methods=['GET'])
def requirement_detail(req_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))

    try:
        with get_db_cursor() as (conn, cur):
            # assigned_to stored as CSV usernames now
            cur.execute("""
                SELECT r.*
                FROM requirements r
                WHERE r.id = %s
            """, (req_id,))
            requirement = cur.fetchone()

            if not requirement:
                flash('Requirement not found', 'danger')
                return redirect(url_for('requirements'))

            is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.args.get('ajax') == '1'
            if is_ajax:
                return render_template('requirement_detail_partial.html', requirement=requirement)
            else:
                return render_template('requirement_detail.html', requirement=requirement)

    except Exception:
        app.logger.exception("Error fetching requirement detail")
        flash('Error loading requirement detail', 'danger')
        return redirect(url_for('requirements'))


# Safety route: handle accidental requests without an ID
@app.route('/update_requirement', methods=['GET', 'POST'])
def update_requirement_root():
    flash('No requirement selected to edit. Please choose a requirement from the list.', 'warning')
    return redirect(url_for('requirements'))


@app.route('/update_requirement/<int:req_id>', methods=['GET', 'POST'])
def update_requirement(req_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))

    try:
        with get_db_cursor() as (conn, cur):
            # Need recruiters/admins for "Assign To" checklist on both GET and POST (when errors)
            cur.execute("SELECT username FROM users WHERE role IN ('recruiter','admin') ORDER BY username ASC")
            recruiters = [r['username'] for r in cur.fetchall()]

            if request.method == 'POST':
                data, errors = validate_requirement_form(request.form)
                if errors:
                    for e in errors:
                        flash(e, 'danger')
                    cur.execute("SELECT * FROM requirements WHERE id = %s", (req_id,))
                    requirement = cur.fetchone()
                    # prepare assigned_usernames for template
                    assigned_usernames = []
                    if requirement and requirement.get('assigned_to'):
                        assigned_usernames = [u.strip() for u in requirement['assigned_to'].split(',') if u.strip()]
                    return render_template('update_requirement.html', requirement=requirement, recruiters=recruiters, assigned_usernames=assigned_usernames)

                cur.execute("""
                    UPDATE requirements SET
                        client_name = %s,
                        requirement_name = %s,
                        experience = %s,
                        mandatory_skills = %s,
                        job_locations = %s,
                        remote = %s,
                        budget = %s,
                        job_description = %s,
                        job_d_th_d = %s,
                        client_linkedin_profile = %s,
                        client_brief_description = %s,
                        assigned_to = %s,
                        status = %s
                    WHERE id = %s
                """, (
                    data['client_name'],
                    data['requirement_name'],
                    data['experience'],
                    data['mandatory_skills'],
                    data['job_locations'],
                    data['remote'],
                    data['budget'],
                    data['job_description'],
                    data['job_d_th_d'],
                    data['client_linkedin_profile'],
                    data['client_brief_description'],
                    data['assigned_to'],  # CSV usernames
                    data['status'],
                    req_id
                ))
                
                conn.commit()
                flash('Requirement updated successfully!', 'success')

                cur.execute("SELECT * FROM requirements WHERE id = %s", (req_id,))
                req = cur.fetchone()
                assigned_users = [u.strip() for u in (req['assigned_to'] or '').split(',') if u.strip()]
                send_requirement_email(req, assigned_users)

                return redirect(url_for('requirement_detail', req_id=req_id))

            else:
                cur.execute("SELECT * FROM requirements WHERE id = %s", (req_id,))
                requirement = cur.fetchone()
                if not requirement:
                    flash('Requirement not found', 'danger')
                    return redirect(url_for('requirements'))
                # prepare assigned_usernames list for pre-checking checkboxes
                assigned_usernames = []
                if requirement.get('assigned_to'):
                    assigned_usernames = [u.strip() for u in requirement['assigned_to'].split(',') if u.strip()]
                return render_template('update_requirement.html', requirement=requirement, recruiters=recruiters, assigned_usernames=assigned_usernames)

    except Exception:
        app.logger.exception("Error updating requirement")
        flash('Error updating requirement', 'danger')
        return redirect(url_for('requirements'))


# -----------------------
# Candidates: List / Add / Edit / View / Delete
# -----------------------
@app.route('/requirement/<int:req_id>/candidates')
def requirement_candidates(req_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))

    # pagination + search + sort
    page, per_page = sanitize_page_params(
        request.args.get("page"),
        request.args.get("per_page"),
        default_per_page=50,
        max_per_page=200
    )

    # multiple search fields
    name = (request.args.get('name') or '').strip()
    phone = (request.args.get('phone') or '').strip()
    email = (request.args.get('email') or '').strip()
    location = (request.args.get('location') or '').strip()

    # NEW filters
    calling_status = (request.args.get('calling_status') or '').strip()
    profile_status = (request.args.get('profile_status') or '').strip()


    sort_by = request.args.get('sort_by', 'added_date')
    sort_dir = 'DESC' if request.args.get('sort_dir', 'desc') == 'desc' else 'ASC'
    offset = (page - 1) * per_page

    try:
        with get_db_cursor() as (conn, cur):
            # requirement existence check
            cur.execute("SELECT id, client_name, requirement_name, assigned_to FROM requirements WHERE id = %s", (req_id,))
            req = cur.fetchone()
            if not req:
                flash('Requirement not found', 'danger')
                return redirect(url_for('requirements'))

            # recruiter: ensure they can view candidates for requirements assigned to them
            if session.get('role') == 'recruiter':
                me = session.get('username')
                if not me:
                    cur.execute("SELECT username FROM users WHERE id = %s", (session['user_id'],))
                    r = cur.fetchone()
                    me = r.get('username') if r else None
                    session['username'] = me
                if me:
                    assigned_csv = (req.get('assigned_to') or '').lower()
                    if me.lower() not in assigned_csv:
                        flash('You do not have access to candidates for this requirement', 'danger')
                        return redirect(url_for('requirements'))

            base = "SELECT * FROM candidates WHERE requirement_id = %s"
            params = [req_id]

            if name:
                base += " AND candidate_name ILIKE %s"
                params.append(f"%{name}%")
            if phone:
                base += " AND COALESCE(phones::text, '') ILIKE %s"
                params.append(f"%{phone}%")
            if email:
                base += " AND COALESCE(emails::text, '') ILIKE %s"
                params.append(f"%{email}%")
            if location:
                base += " AND current_location ILIKE %s"
                params.append(f"%{location}%")


            # NEW: calling_status/profile_status filters
            if calling_status:
                base += " AND calling_status = %s"
                params.append(calling_status)
            if profile_status:
                base += " AND profile_status = %s"
                params.append(profile_status)
            allowed_sort = {'added_date': 'added_date', 'candidate_name': 'candidate_name', 'application_date': 'application_date'}
            order_col = allowed_sort.get(sort_by, 'added_date')

            q = f"{base} ORDER BY {order_col} {sort_dir} LIMIT %s OFFSET %s"
            cur.execute(q, tuple(params + [per_page, offset]))
            candidates = cur.fetchall()

            # Normalize phones/emails to Python lists so templates can rely on lists
            for c in candidates:
                c['phones'] = normalize_list_field(c.get('phones'))
                c['emails'] = normalize_list_field(c.get('emails'))

            # count total
            cnt_q = "SELECT COUNT(*) as total FROM candidates WHERE requirement_id = %s"
            cnt_params = [req_id]
            if name:
                cnt_q += " AND candidate_name ILIKE %s"
                cnt_params.append(f"%{name}%")
            if phone:
                cnt_q += " AND COALESCE(phones::text, '') ILIKE %s"
                cnt_params.append(f"%{phone}%")
            if email:
                cnt_q += " AND COALESCE(emails::text, '') ILIKE %s"
                cnt_params.append(f"%{email}%")
            if location:
                cnt_q += " AND current_location ILIKE %s"
                cnt_params.append(f"%{location}%")
            if calling_status:
                cnt_q += " AND calling_status = %s"
                cnt_params.append(calling_status)
            if profile_status:
                cnt_q += " AND profile_status = %s"
                cnt_params.append(profile_status)
            cur.execute(cnt_q, tuple(cnt_params))
            row = cur.fetchone()
            total = row['total'] if row else 0   # ✅ fixed


    except Exception:
        app.logger.exception("Error loading candidates")
        flash("Error loading candidates", "danger")
        candidates = []
        total = 0
        req = {'id': req_id, 'client_name': '', 'requirement_name': ''}

    paginator = Paginator(
        total=total,
        page=page,
        per_page=per_page,
        base_url=url_for("requirement_candidates", req_id=req_id),
        args=request.args
    )

    return render_template(
        'candidates_list.html',
        requirement=req,
        candidates=candidates,
        paginator=paginator,
        name=name, phone=phone, email=email, location=location,
                calling_status=calling_status, profile_status=profile_status,
        sort_by=sort_by, sort_dir=sort_dir
    )

@app.route('/requirement/<int:req_id>/candidates/template.xlsx')
def download_candidate_template(req_id):
    from io import BytesIO
    from flask import send_file
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(SHEET_HEADERS)  # Using the same headers list we defined earlier

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    filename = f"candidate_template_req_{req_id}.xlsx"
    return send_file(
        bio,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@app.route('/requirement/<int:req_id>/candidates/add', methods=['GET', 'POST'])
def add_candidate(req_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))

    try:
        with get_db_cursor() as (conn, cur):
            cur.execute("SELECT id, requirement_name, assigned_to FROM requirements WHERE id = %s", (req_id,))
            req = cur.fetchone()
            if not req:
                flash('Requirement not found', 'danger')
                return redirect(url_for('requirements'))

            # recruiter permission check
            if session.get('role') == 'recruiter':
                me = session.get('username')
                assigned_csv = (req.get('assigned_to') or '').lower()
                if me and me.lower() not in assigned_csv:
                    flash('You do not have access to add candidates for this requirement', 'danger')
                    return redirect(url_for('requirements'))

            if request.method == 'POST':
                data, errors = validate_candidate_form(request.form)
                if errors:
                    for e in errors:
                        flash(e, 'danger')
                    return render_template('candidate_form.html', requirement=req, candidate=None, data=request.form)

                # Handle multiple phones/emails → store as JSON (use psycopg2's Json adapter)
                phones_list = request.form.getlist('phones')
                emails_list = request.form.getlist('emails')

                cur.execute("""
                    INSERT INTO candidates (
                        requirement_id, application_date, job_title, candidate_name, current_company,
                        total_experience, phones, emails, notice_period, current_location,
                        preferred_locations, ctc_current, ectc, key_skills, education,
                        post_graduation, pf_docs_confirm, notice_period_details,
                        current_ctc_lpa, expected_ctc_lpa, employee_size, companies_worked,
                        calling_status, profile_status, comments, interview_date, interview_time, added_by, added_date, updated_date
                    ) VALUES (
                        %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,
                        %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,
                        %s,%s,%s,%s,%s, %s,%s, %s, now(), now()
                    )
                """, (
                    req_id,
                    data['application_date'],
                    data['job_title'],
                    data['candidate_name'],
                    data['current_company'],
                    data['total_experience'],
                    psycopg2.extras.Json(phones_list),
                    psycopg2.extras.Json(emails_list),
                    data['notice_period'],
                    data['current_location'],
                    data['preferred_locations'],
                    data['ctc_current'],
                    data['ectc'],
                    data['key_skills'],
                    data['education'],
                    data['post_graduation'],
                    data['pf_docs_confirm'],
                    data['notice_period_details'],
                    data['current_ctc_lpa'],
                    data['expected_ctc_lpa'],
                    data['employee_size'],
                    data['companies_worked'],
                    data['calling_status'],
                    data['profile_status'],
                    data['comments'],
                            data.get('interview_date'),
                    data.get('interview_time'),
                    session.get('username') or 'system'
                ))
                conn.commit()
                flash('Candidate added successfully!', 'success')
                return redirect(url_for('requirement_candidates', req_id=req_id))

            return render_template('candidate_form.html', requirement=req, candidate=None, data={})

    except Exception:
        app.logger.exception("Error adding candidate")
        flash('Error adding candidate', 'danger')
        return redirect(url_for('requirement_candidates', req_id=req_id))


@app.route('/candidate/<int:cand_id>/edit', methods=['GET', 'POST'])
def edit_candidate(cand_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))

    try:
        with get_db_cursor() as (conn, cur):
            cur.execute("SELECT * FROM candidates WHERE id = %s", (cand_id,))
            cand = cur.fetchone()
            if not cand:
                flash('Candidate not found', 'danger')
                return redirect(url_for('requirements'))

            # requirement for breadcrumb and permission check
            cur.execute("SELECT id, requirement_name, assigned_to FROM requirements WHERE id = %s", (cand['requirement_id'],))
            req = cur.fetchone()
            if not req:
                flash('Requirement not found', 'danger')
                return redirect(url_for('requirements'))

            if session.get('role') == 'recruiter':
                me = session.get('username')
                assigned_csv = (req.get('assigned_to') or '').lower()
                if me and me.lower() not in assigned_csv:
                    flash('You do not have access to edit candidates for this requirement', 'danger')
                    return redirect(url_for('requirements'))

            if request.method == 'POST':
                data, errors = validate_candidate_form(request.form)
                if errors:
                    for e in errors:
                        flash(e, 'danger')
                    return render_template('candidate_form.html', requirement=req, candidate=cand, data=request.form)

                cur.execute("""
                    UPDATE candidates SET
                      application_date=%s, job_title=%s, candidate_name=%s, current_company=%s,
                      total_experience=%s, phones=%s, emails=%s, notice_period=%s, current_location=%s,
                      preferred_locations=%s, ctc_current=%s, ectc=%s, key_skills=%s, education=%s,
                      post_graduation=%s, pf_docs_confirm=%s, notice_period_details=%s,
                      current_ctc_lpa=%s, expected_ctc_lpa=%s, employee_size=%s, companies_worked=%s,
                      calling_status=%s, profile_status=%s, comments=%s, interview_date=%s, interview_time=%s, updated_date=now()
                    WHERE id=%s
                """, (
                    data.get('application_date'), data.get('job_title'), data.get('candidate_name'), data.get('current_company'),
                    data.get('total_experience'), psycopg2.extras.Json(data.get('phones') or []), psycopg2.extras.Json(data.get('emails') or []),
                    data.get('notice_period'), data.get('current_location'), data.get('preferred_locations'),
                    data.get('ctc_current'), data.get('ectc'), data.get('key_skills'), data.get('education'),
                    data.get('post_graduation'), data.get('pf_docs_confirm'), data.get('notice_period_details'),
                    data.get('current_ctc_lpa'), data.get('expected_ctc_lpa'), data.get('employee_size'),
                    data.get('companies_worked'), data.get('calling_status'), data.get('profile_status'), data.get('comments'),
                        data.get('interview_date'), data.get('interview_time'),
                    cand_id
                ))
                conn.commit()
                flash('Candidate updated successfully!', 'success')
                return redirect(url_for('requirement_candidates', req_id=cand['requirement_id']))

            # GET: ensure JSON fields are python lists
            cand['phones'] = normalize_list_field(cand.get('phones'))
            cand['emails'] = normalize_list_field(cand.get('emails'))
            return render_template('candidate_form.html', requirement=req, candidate=cand, data=cand)
    except Exception:
        app.logger.exception("Error editing candidate")
        flash('Error editing candidate', 'danger')
        return redirect(url_for('requirements'))


@app.route('/candidate/<int:cand_id>/delete', methods=['POST'])
def delete_candidate(cand_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    try:
        with get_db_cursor() as (conn, cur):
            cur.execute("SELECT requirement_id FROM candidates WHERE id = %s", (cand_id,))
            row = cur.fetchone()
            if not row:
                flash('Candidate not found', 'danger')
                return redirect(url_for('requirements'))
            req_id = row['requirement_id']
            cur.execute("DELETE FROM candidates WHERE id = %s", (cand_id,))
            conn.commit()
            flash('Candidate deleted successfully!', 'success')
            return redirect(url_for('requirement_candidates', req_id=req_id))
    except Exception:
        app.logger.exception("Error deleting candidate")
        flash('Error deleting candidate', 'danger')
        return redirect(url_for('requirements'))


@app.route('/candidate/<int:cand_id>')
def view_candidate(cand_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    try:
        with get_db_cursor() as (conn, cur):
            cur.execute("SELECT c.*, r.client_name, r.requirement_name, r.assigned_to FROM candidates c LEFT JOIN requirements r ON c.requirement_id=r.id WHERE c.id = %s", (cand_id,))
            cand = cur.fetchone()
            if not cand:
                flash('Candidate not found', 'danger')
                return redirect(url_for('requirements'))

            # recruiter permission check
            if session.get('role') == 'recruiter':
                me = session.get('username')
                assigned_csv = (cand.get('assigned_to') or '').lower()
                if me and me.lower() not in assigned_csv:
                    flash('You do not have access to view this candidate', 'danger')
                    return redirect(url_for('requirements'))

            # Normalize phones / emails to lists
            cand['phones'] = normalize_list_field(cand.get('phones'))
            cand['emails'] = normalize_list_field(cand.get('emails'))
            return render_template('candidate_detail.html', candidate=cand)
    except Exception:
        app.logger.exception("Error loading candidate detail")
        flash('Error loading candidate detail', 'danger')
        return redirect(url_for('requirements'))


# New: partial candidate detail used for modal (AJAX)
@app.route('/candidate/<int:cand_id>/partial')
def candidate_partial(cand_id):
    if 'user_id' not in session:
        return jsonify({'error': 'unauthenticated'}), 401
    try:
        with get_db_cursor() as (conn, cur):
            cur.execute("SELECT c.*, r.client_name, r.requirement_name, r.assigned_to FROM candidates c LEFT JOIN requirements r ON c.requirement_id=r.id WHERE c.id = %s", (cand_id,))
            cand = cur.fetchone()
            if not cand:
                return ("<div class='p-3'>Candidate not found</div>", 404)

            # recruiter permission check
            if session.get('role') == 'recruiter':
                me = session.get('username')
                assigned_csv = (cand.get('assigned_to') or '').lower()
                if me and me.lower() not in assigned_csv:
                    return ("<div class='p-3'>You do not have access to view this candidate.</div>", 403)

            cand['phones'] = normalize_list_field(cand.get('phones'))
            cand['emails'] = normalize_list_field(cand.get('emails'))
            return render_template('candidate_detail_partial.html', candidate=cand)
    except Exception:
        app.logger.exception("Error loading candidate partial")
        return ("<div class='p-3'>Error loading candidate detail</div>", 500)




@app.route('/candidate/<int:cand_id>/status', methods=['POST'])
def update_candidate_status(cand_id):
    if 'user_id' not in session:
        return jsonify({'ok': False, 'error': 'unauthenticated'}), 401
    try:
        data = request.get_json(silent=True) or {}
        calling_status = (data.get('calling_status') or '').strip()
        profile_status = (data.get('profile_status') or '').strip()
        interview_date = (data.get('interview_date') or '').strip()
        interview_time = (data.get('interview_time') or '').strip()
        with get_db_cursor() as (conn, cur):
            sets = []
            params = []
            if calling_status != '':
                sets.append("calling_status = %s")
                params.append(calling_status)
            if profile_status != '':
                sets.append("profile_status = %s")
                params.append(profile_status)
            # empty -> NULL to clear
            sets.append("interview_date = %s"); params.append(interview_date if interview_date else None)
            sets.append("interview_time = %s"); params.append(interview_time if interview_time else None)
            sets.append("updated_date = now()")
            sql = "UPDATE candidates SET " + ", ".join(sets) + " WHERE id = %s"
            params.append(cand_id)
            cur.execute(sql, tuple(params))
            conn.commit()
            return jsonify({'ok': True})
    except Exception:
        app.logger.exception("Error updating candidate status inline")
        return jsonify({'ok': False, 'error': 'server_error'}), 500
# Export route: produces XLSX of filtered results or selected ids
@app.route('/requirement/<int:req_id>/candidates/export')
def export_candidates(req_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))

    # fetch params (same as list)
    name = (request.args.get('name') or '').strip()
    phone = (request.args.get('phone') or '').strip()
    email = (request.args.get('email') or '').strip()
    location = (request.args.get('location') or '').strip()
    ids = (request.args.get('ids') or '').strip()  # comma separated ids (optional)

    try:
        with get_db_cursor() as (conn, cur):
            # requirement existence check
            cur.execute("SELECT id, requirement_name FROM requirements WHERE id = %s", (req_id,))
            req = cur.fetchone()
            if not req:
                flash('Requirement not found', 'danger')
                return redirect(url_for('requirements'))

            base = "SELECT * FROM candidates WHERE requirement_id = %s"
            params = [req_id]

            if ids:
                # prefer selected ids if provided
                try:
                    id_list = [int(x) for x in ids.split(',') if x.strip()]
                    if id_list:
                        base += " AND id = ANY(%s)"
                        params.append(tuple(id_list))
                except Exception:
                    # fallback to not using ids if parse fails
                    pass
            else:
                # apply filters only when ids not provided
                if name:
                    base += " AND candidate_name ILIKE %s"
                    params.append(f"%{name}%")
                if phone:
                    base += " AND COALESCE(phones::text, '') ILIKE %s"
                    params.append(f"%{phone}%")
                if email:
                    base += " AND COALESCE(emails::text, '') ILIKE %s"
                    params.append(f"%{email}%")
                if location:
                    base += " AND current_location ILIKE %s"
                    params.append(f"%{location}%")

            q = f"{base} ORDER BY added_date DESC"
            cur.execute(q, tuple(params))
            rows = cur.fetchall()

            # prepare workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Candidates"

            # If no rows, still create workbook with a message/header
            if not rows:
                ws.append(["No candidates found for the selected filters"])
            else:
                # use keys from first row as headers (preserve order by converting to list)
                first = rows[0]
                headers = list(first.keys())
                # Write header row
                ws.append(headers)
                for r in rows:
                    row_values = []
                    for h in headers:
                        v = r.get(h)
                        if isinstance(v, (list, dict)):
                            try:
                                row_values.append(json.dumps(v, ensure_ascii=False))
                            except Exception:
                                row_values.append(str(v))
                        elif isinstance(v, datetime):
                            row_values.append(v.strftime('%Y-%m-%d %H:%M:%S'))
                        else:
                            row_values.append("" if v is None else v)
                    ws.append(row_values)

            # prepare file for download
            file_stream = io.BytesIO()
            wb.save(file_stream)
            file_stream.seek(0)

            filename = f"candidates_req_{req_id}.xlsx"
            return send_file(
                file_stream,
                as_attachment=True,
                download_name=filename,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
    except Exception:
        app.logger.exception("Error exporting candidates")
        flash('Error exporting candidates', 'danger')
        return redirect(url_for('requirement_candidates', req_id=req_id))


# --- Delete / forgot / reset / dashboard_data unchanged from earlier file ---

# =====================
# Bulk Import (Excel + Paste) endpoints
# =====================

@app.route('/requirement/<int:req_id>/candidates/import/upload', methods=['POST'])
def import_candidates_upload(req_id):
    if 'user_id' not in session:
        return jsonify({'error': 'unauthenticated'}), 401
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400
        f = request.files['file']
        filename = f.filename or ''
        if not filename.lower().endswith('.xlsx'):
            return jsonify({'error': 'Only .xlsx files are supported'}), 400
        import openpyxl, io
        file_bytes = f.read()
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb.active
        headers = [ (c.value if c.value is not None else '') for c in ws[1] ]
        idx_map = _smart_map_headers(headers)

        rows = []
        row_errors = {}
        for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if all((cell.value is None or str(cell.value).strip()=='' ) for cell in row):
                continue
            mapped = {}
            for idx, cell in enumerate(row):
                key = idx_map.get(idx, f'col_{idx+1}')
                mapped[key] = cell.value
            norm = _normalize_row(mapped)
            errs = _validate_row_smart(norm)
            if errs:
                row_errors[i] = errs
            rows.append(norm)
        return jsonify({'rows': rows, 'row_errors': row_errors})
    except Exception as e:
        app.logger.exception("Upload parse error: %s", e)
        return jsonify({'error': 'Failed to read Excel'}), 500




@app.route('/delete_requirement/<int:req_id>', methods=['POST'])
def delete_requirement(req_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))

    try:
        with get_db_cursor() as (conn, cur):
            cur.execute("DELETE FROM requirements WHERE id = %s", (req_id,))
            conn.commit()
            flash('Requirement deleted successfully!', 'success')
    except Exception:
        app.logger.exception("Error deleting requirement")
        flash('Error deleting requirement', 'danger')

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.args.get('ajax') == '1':
        return jsonify({'status': 'ok', 'deleted_id': req_id})

    return redirect(url_for('requirements'))


@app.route('/forgot_password', methods=['GET', 'POST'])
def forgot_password():
    if request.method == 'POST':
        email = request.form.get('email', '').strip()
        if not email:
            flash('Please provide your email address', 'danger')
            return redirect(url_for('forgot_password'))

        try:
            with get_db_cursor() as (conn, cur):
                cur.execute("SELECT id, email FROM users WHERE email = %s", (email,))
                user = cur.fetchone()
                if user:
                    token = ts.dumps(email, salt='password-reset-salt')
                    reset_url = url_for('reset_password', token=token, _external=True)
                    app.logger.info(f'Password reset for {email}: {reset_url}')
                flash('If the email exists in our system you will receive a password reset link shortly.', 'info')
                return redirect(url_for('login'))
        except Exception:
            app.logger.exception("Error in forgot_password")
            flash('Error processing request', 'danger')
            return redirect(url_for('forgot_password'))

    return render_template('forgot_password.html')


@app.route('/reset_password/<token>', methods=['GET', 'POST'])
def reset_password(token):
    try:
        email = ts.loads(token, salt='password-reset-salt', max_age=3600)
    except Exception:
        flash('Invalid or expired token', 'danger')
        return redirect(url_for('login'))

    if request.method == 'POST':
        password = request.form.get('password', '')
        password2 = request.form.get('password2', '')
        if not password or password != password2:
            flash('Passwords do not match', 'danger')
            return redirect(url_for('reset_password', token=token))
        try:
            with get_db_cursor() as (conn, cur):
                hashed = generate_password_hash(password)
                cur.execute("UPDATE users SET password_hash = %s WHERE email = %s", (hashed, email))
                conn.commit()
                flash('Password reset successful. Please log in.', 'success')
                return redirect(url_for('login'))
        except Exception:
            app.logger.exception("Error resetting password")
            flash('Error resetting password', 'danger')
            return redirect(url_for('login'))

    return render_template('reset_password.html', token=token)




@app.route('/users')
def users():
    if 'user_id' not in session or session.get('role') != 'admin':
        flash('Access denied', 'danger')
        return redirect(url_for('dashboard'))

    search = request.args.get('search', '').strip()
    query = "SELECT id, first_name, last_name, email, role, status FROM users"
    params = []
    if search:
        query += " WHERE first_name ILIKE %s OR last_name ILIKE %s OR email ILIKE %s"
        params.extend([f"%{search}%", f"%{search}%", f"%{search}%"])
    query += " ORDER BY id ASC"

    with get_db_cursor() as (conn, cur):
        cur.execute(query, tuple(params))
        users_list = cur.fetchall()

    return render_template('users.html', users=users_list, search=search)


@app.route('/add_user', methods=['GET', 'POST'])
def add_user():
    if 'user_id' not in session or session.get('role') != 'admin':
        flash('Access denied', 'danger')
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        first_name = request.form['first_name']
        last_name = request.form['last_name']
        email = request.form['email']
        password = request.form['password']
        role = request.form['role']
        status = request.form['status']

        # ✅ Auto-generate username from email (before @)
        username = email.split('@')[0].strip().lower()

        hashed_password = generate_password_hash(password)

        with get_db_cursor() as (conn, cur):
            cur.execute("""
                INSERT INTO users (username, first_name, last_name, email, password_hash, role, status)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """, (username, first_name, last_name, email, hashed_password, role, status))
            conn.commit()

        flash('User added successfully', 'success')
        return redirect(url_for('users'))

    return render_template('add_user.html')


@app.route('/edit_user/<int:user_id>', methods=['GET', 'POST'])
def edit_user(user_id):
    if 'user_id' not in session or session.get('role') != 'admin':
        flash('Access denied', 'danger')
        return redirect(url_for('dashboard'))

    with get_db_cursor() as (conn, cur):
        cur.execute("SELECT id, first_name, last_name, email, role, status FROM users WHERE id=%s", (user_id,))
        user = cur.fetchone()

    if not user:
        flash('User not found', 'danger')
        return redirect(url_for('users'))

    if request.method == 'POST':
        first_name = request.form['first_name']
        last_name = request.form['last_name']
        email = request.form['email']
        role = request.form['role']
        status = request.form['status']

        with get_db_cursor() as (conn, cur):
            cur.execute("UPDATE users SET first_name=%s, last_name=%s, email=%s, role=%s, status=%s WHERE id=%s",
                        (first_name, last_name, email, role, status, user_id))
            conn.commit()

        flash('User updated successfully', 'success')
        return redirect(url_for('users'))

    return render_template('edit_user.html', user=user)


@app.route('/delete_user/<int:user_id>', methods=['POST'])
def delete_user(user_id):
    if 'user_id' not in session or session.get('role') != 'admin':
        flash('Access denied', 'danger')
        return redirect(url_for('dashboard'))

    with get_db_cursor() as (conn, cur):
        cur.execute("DELETE FROM users WHERE id=%s", (user_id,))
        conn.commit()

    flash('User deleted successfully', 'success')
    return redirect(url_for('users'))


@app.route('/reset_user_password/<int:user_id>', methods=['GET', 'POST'])
def reset_user_password(user_id):
    if 'user_id' not in session or session.get('role') != 'admin':
        flash('Access denied', 'danger')
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        new_password = request.form['password']
        hashed_password = generate_password_hash(new_password)

        with get_db_cursor() as (conn, cur):
            cur.execute("UPDATE users SET password_hash=%s WHERE id=%s", (hashed_password, user_id))
            conn.commit()

        flash('Password reset successfully', 'success')
        return redirect(url_for('users'))

    return render_template('reset_user_password.html', user_id=user_id)

print('==== ROUTE MAP START ====')
for rule in app.url_map.iter_rules():
    print(rule, '->', rule.endpoint)
print('==== ROUTE MAP END ====')

@app.route('/change_password', methods=['GET', 'POST'])
def change_password():
    if 'user_id' not in session:
        flash('Please log in to change your password', 'warning')
        return redirect(url_for('login'))

    if request.method == 'POST':
        current_password = request.form.get('current_password', '')
        new_password = request.form.get('new_password', '')
        confirm_password = request.form.get('confirm_password', '')

        if not current_password or not new_password or not confirm_password:
            flash('Please fill in all fields', 'danger')
            return redirect(url_for('change_password'))

        if new_password != confirm_password:
            flash('New passwords do not match', 'danger')
            return redirect(url_for('change_password'))

        try:
            with get_db_cursor() as (conn, cur):
                cur.execute("SELECT password_hash FROM users WHERE id = %s", (session['user_id'],))
                row = cur.fetchone()
                if not row or not row.get('password_hash'):
                    flash('Account not found', 'danger')
                    return redirect(url_for('change_password'))
                if not check_password_hash(row['password_hash'], current_password):
                    flash('Current password is incorrect', 'danger')
                    return redirect(url_for('change_password'))

                hashed = generate_password_hash(new_password)
                cur.execute("UPDATE users SET password_hash = %s WHERE id = %s", (hashed, session['user_id']))
                conn.commit()
                flash('Password updated successfully.', 'success')
                return redirect(url_for('dashboard'))
        except Exception:
            app.logger.exception('Error changing password')
            flash('Error changing password', 'danger')
            return redirect(url_for('change_password'))

    return render_template('change_password.html')

if __name__ == '__main__':
    app.run(debug=True)





@app.route('/requirement/<int:req_id>/candidates/import/commit', methods=['POST'])
def import_candidates_commit(req_id):
    if 'user_id' not in session:
        return jsonify({'error': 'unauthenticated'}), 401
    payload = request.get_json(silent=True) or {}
    rows = payload.get('rows') or []
    if not isinstance(rows, list) or not rows:
        return jsonify({'error': 'No rows to import'}), 400
    inserted = 0
    skipped = []
    try:
        with get_db_cursor() as (conn, cur):
            for idx, r in enumerate(rows, start=2):
                errs = _validate_row_smart(r)
                if errs:
                    skipped.append({'row': idx, 'reasons': errs, 'data': r})
                    continue
                try:
                    cur.execute("""
                        INSERT INTO candidates (
                            requirement_id, application_date, job_title, candidate_name, current_company,
                            total_experience, phones, emails, notice_period, current_location,
                            preferred_locations, ctc_current, ectc, key_skills, education,
                            post_graduation, pf_docs_confirm, notice_period_details,
                            current_ctc_lpa, expected_ctc_lpa, employee_size, companies_worked,
                            calling_status, profile_status, comments, interview_date, interview_time, added_by, added_date, updated_date
                        ) VALUES (
                            %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,
                            %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,
                            %s,%s,%s,%s,%s, %s,%s, %s, now(), now()
                        )
                    """, (
                        req_id,
                        r.get('application_date',''),
                        r.get('job_title',''),
                        r.get('candidate_name',''),
                        r.get('current_company',''),
                        r.get('total_experience',''),
                        r.get('phones',''),
                        r.get('emails',''),
                        r.get('notice_period',''),
                        r.get('current_location',''),
                        r.get('preferred_locations',''),
                        r.get('ctc_current',''),
                        r.get('ectc',''),
                        r.get('key_skills',''),
                        r.get('education',''),
                        r.get('post_graduation',''),
                        r.get('pf_docs_confirm',''),
                        r.get('notice_period_details',''),
                        r.get('current_ctc_lpa',''),
                        r.get('expected_ctc_lpa',''),
                        r.get('employee_size',''),
                        r.get('companies_worked',''),
                        r.get('calling_status',''),
                        r.get('profile_status',''),
                        r.get('comments',''),
                        session.get('user_id') or session.get('username') or 'system'
                    ))
                    inserted += 1
                except Exception as ie:
                    app.logger.exception("Insert failed for row %s: %s", idx, ie)
                    skipped.append({'row': idx, 'reasons': ['DB insert failed'], 'data': r})
            conn.commit()
        return jsonify({'inserted': inserted, 'skipped': skipped})
    except Exception as e:
        app.logger.exception("Bulk import commit error: %s", e)
        return jsonify({'error': 'Server error while saving'}), 500




@app.route('/requirement/<int:req_id>/candidates/paste/preview', methods=['POST'])
def paste_candidates_preview(req_id):
    if 'user_id' not in session:
        return jsonify({'error': 'unauthenticated'}), 401
    payload = request.get_json(silent=True) or {}
    text = payload.get('text', '')
    if not str(text).strip():
        return jsonify({'error': 'No data pasted'}), 400
    lines = [ln for ln in text.splitlines() if ln.strip()]
    if not lines:
        return jsonify({'error': 'No rows found'}), 400

    delim = '	' if any('	' in ln for ln in lines) else (',' if any(',' in ln for ln in lines) else '|')
    header_cells = [c.strip() for c in lines[0].split(delim)]
    header_map = _smart_map_headers(header_cells)
    header_match_score = sum(1 for i,_ in enumerate(header_cells) if header_map.get(i,'').strip() in SHEET_COLUMNS)
    has_header = header_match_score >= max(2, len(header_cells)//2)

    rows = []
    row_errors = {}
    start_idx = 1 if has_header else 0
    for i, ln in enumerate(lines[start_idx:], start=(2 if has_header else 1)):
        cells = [c.strip() for c in ln.split(delim)]
        mapped = {}
        if has_header:
            for idx, val in enumerate(cells):
                key = header_map.get(idx, f'col_{idx+1}')
                mapped[key] = val
        else:
            for idx, val in enumerate(cells):
                mapped[f'col_{idx+1}'] = val
        norm = _normalize_row(mapped)
        errs = _validate_row_smart(norm)
        if errs:
            row_errors[i] = errs
        rows.append(norm)

    return jsonify({'rows': rows, 'row_errors': row_errors})


@app.route('/requirement/<int:req_id>/candidates/paste/commit', methods=['POST'])
def paste_candidates_commit(req_id):
    return import_candidates_commit(req_id)






@app.route('/candidate/new', methods=['GET','POST'])
def candidate_new():
    """Redirect helper to add_candidate which expects a requirement id (req_id).
    Call this route with ?req_id=123 otherwise user is redirected to the requirements list.
    """
    req_id = request.args.get('req_id')
    if not req_id:
        flash('No requirement selected to add candidate. Please open the requirement and click Add Candidate.', 'warning')
        return redirect(url_for('requirements'))
    try:
        return redirect(url_for('add_candidate', req_id=int(req_id)))
    except Exception:
        flash('Invalid requirement id', 'danger')
        return redirect(url_for('requirements'))





# API compatibility wrappers for Import Wizard frontend




# ----------------------
# Import Wizard routes
# ----------------------

SHEET_COLUMNS = [
    "application_date","job_title","candidate_name","current_company","total_experience",
    "phones","emails","notice_period","current_location","preferred_locations","ctc_current",
    "ectc","key_skills","education","post_graduation","pf_docs_confirm","notice_period_details",
    "current_ctc_lpa","expected_ctc_lpa","employee_size","companies_worked",
    "calling_status","profile_status","comments"
]

ALIASES = {
    "candidate name": "candidate_name",
    "name": "candidate_name",
    "mobile": "phones",
    "phone": "phones",
    "email": "emails",
    "current ctc": "current_ctc_lpa",
    "expected ctc": "expected_ctc_lpa",
    "location": "current_location",
    "application date": "application_date",
    "job": "job_title",
}

def _smart_map_headers(headers):
    suggested = {}
    unmapped = []
    for h in headers:
        key = (h or "").strip()
        kclean = key.lower().replace("-", " ").replace("_", " ").strip()
        sys = ALIASES.get(kclean)
        if not sys and kclean in SHEET_COLUMNS:
            sys = kclean
        if not sys:
            unmapped.append(key)
        suggested[key] = sys or ""
    return suggested, unmapped

import io, csv, datetime
from flask import request, jsonify, render_template
from flask_login import login_required

def _parse_rows_from_csv(text):
    f = io.StringIO(text)
    sniffer = csv.Sniffer()
    try:
        dialect = sniffer.sniff(text.splitlines()[0])
    except Exception:
        dialect = csv.excel
    reader = csv.DictReader(f, dialect=dialect)
    headers = reader.fieldnames or []
    rows_raw = list(reader)
    samples = {h: [] for h in headers}
    for r in rows_raw[:5]:
        for h in headers:
            v = (r.get(h) or "").strip()
            if v:
                samples[h].append(v)
    suggested, unmapped = _smart_map_headers(headers)
    mapped = []
    for r in rows_raw:
        out = {k: "" for k in SHEET_COLUMNS}
        for h in headers:
            sys = suggested.get(h) or ""
            if sys:
                out[sys] = (r.get(h) or "").strip()
        mapped.append(out)
    return headers, samples, suggested, unmapped, mapped

@app.route("/requirement/<int:req_id>/candidates/import", methods=["GET"], endpoint="import_wizard")
@login_required
def import_candidates_wizard(req_id):
    try:
        req = get_requirement(req_id)
    except NameError:
        req = {"id": req_id, "client_name": "", "requirement_name": f"Requirement {req_id}"}
    return render_template("import.html", requirement=req, sheet_columns=SHEET_COLUMNS)

@app.route("/api/import/parse", methods=["POST"])
@login_required
def api_import_parse():
    req_id = request.form.get("requirement_id", type=int)
    f = request.files.get("file")
    pasted = request.form.get("text", "")

    if f and f.filename.lower().endswith((".csv", ".tsv", ".txt")):
        text = f.read().decode("utf-8", errors="ignore")
        headers, samples, suggested, unmapped, rows = _parse_rows_from_csv(text)
    elif pasted:
        headers, samples, suggested, unmapped, rows = _parse_rows_from_csv(pasted)
    else:
        return jsonify({"ok": False, "error": "Only CSV/TSV or pasted data supported."})

    return jsonify({
        "ok": True,
        "headers": headers,
        "samples": samples,
        "suggested_mapping": suggested,
        "unmapped_headers": unmapped,
        "rows": rows
    })

@app.route("/api/import/validate", methods=["POST"])
@login_required
def api_import_validate():
    data = request.get_json(force=True) or {}
    rows = data.get("rows", [])
    today = datetime.date.today().isoformat()
    errors = []
    normalized = []
    for i, r in enumerate(rows):
        rr = {k: r.get(k, "") for k in SHEET_COLUMNS}
        if not rr.get("candidate_name"):
            errors.append({"row_index": i, "field": "candidate_name", "message": "Candidate name is required"})
        if not rr.get("application_date"):
            rr["application_date"] = today
        normalized.append(rr)
    return jsonify({"ok": True, "rows": normalized, "errors": errors})


# routes/export.py
from flask import Blueprint, request, send_file, jsonify
import io
import pandas as pd
from datetime import datetime
# from models import Candidate  # <-- import your ORM model

export_bp = Blueprint("export_bp", __name__)

@export_bp.route("/export_candidates", methods=["POST"])
def export_candidates():
    ids = request.json.get("ids", [])
    if not ids:
        return jsonify({"error": "No candidate IDs provided"}), 400

    # Fetch candidates from DB (adjust query for your ORM)
    candidates = Candidate.query.filter(Candidate.id.in_(ids)).all()

    # Build rows
    rows = []
    for c in candidates:
        rows.append({
            "application_date": getattr(c, "application_date", ""),
            "job_title": getattr(c, "job_title", ""),
            "candidate_name": getattr(c, "candidate_name", ""),
            "current_company": getattr(c, "current_company", ""),
            "total_experience": getattr(c, "total_experience", ""),
            "phones": ", ".join(c.phones) if getattr(c, "phones", None) else "",
            "emails": ", ".join(c.emails) if getattr(c, "emails", None) else "",
            "notice_period": getattr(c, "notice_period", ""),
            "current_location": getattr(c, "current_location", ""),
            "preferred_locations": getattr(c, "preferred_locations", ""),
            "ctc_current": getattr(c, "ctc_current", ""),
            "ectc": getattr(c, "ectc", ""),
            "calling_status": getattr(c, "calling_status", ""),
            "profile_status": getattr(c, "profile_status", ""),
            "comments": getattr(c, "comments", ""),
            "added_date": getattr(c, "added_date", ""),
            "updated_date": getattr(c, "updated_date", ""),
            "added_by": getattr(c, "added_by", ""),
        })

    # Convert to DataFrame
    df = pd.DataFrame(rows, columns=[
        "application_date","job_title","candidate_name","current_company",
        "total_experience","phones","emails","notice_period","current_location",
        "preferred_locations","ctc_current","ectc","calling_status","profile_status",
        "comments","added_date","updated_date","added_by"
    ])

    # Save to Excel
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Candidates")
    output.seek(0)

    filename = f"candidates_export_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route("/api/import/save", methods=["POST"])
@login_required
def api_import_save():
    data = request.get_json(force=True) or {}
    req_id = data.get("requirement_id")
    rows = data.get("rows", [])
    inserted = len(rows)
    return jsonify({"ok": True, "saved": inserted})
